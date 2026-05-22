import re
import os
import time
import requests
import feedparser
import cloudscraper
import pandas as pd

from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import Workbook, load_workbook
from playwright.sync_api import sync_playwright

MASTER_FILE = "master_cve.xlsx"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 "
        "(Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 "
        "(KHTML, like Gecko) "
        "Chrome/136 Safari/537.36"
    )
}

CVE_PATTERN = r"CVE[\s\-–—]\d{4}[\s\-–—]\d{4,7}"

scraper = cloudscraper.create_scraper()


def clean_cve(cve):

    return re.sub(
        r"[\s–—\-]+",
        "-",
        cve.upper()
    ).strip("-")


def create_workbook_if_missing():

    if os.path.exists(MASTER_FILE):
        return

    wb = Workbook()

    default_sheet = wb.active

    wb.remove(default_sheet)

    sheets = [

        "BLEEPING",
        "THN",
        "CYBERSECNEWS",
        "SECURITYWEEK",
        "RAPID7",
        "TENABLE",
        "ZDI",
        "ZERODAYCZ",
        "PALOALTO",
        "CISA"

    ]

    for sheet in sheets:

        ws = wb.create_sheet(title=sheet)

        ws.append([
            "CVE",
            "DATE",
            "LINK"
        ])

    wb.save(MASTER_FILE)

def load_existing_keys(sheet_name):

    wb = load_workbook(MASTER_FILE)

    ws = wb[sheet_name]

    keys = set()

    for row in ws.iter_rows(
        min_row=2,
        values_only=True
    ):

        try:

            key = (
                str(row[0]),
                str(row[1]),
                str(row[2])
            )

            keys.add(key)

        except:
            continue

    wb.close()

    return keys

def append_rows(sheet_name, rows):

    wb = load_workbook(MASTER_FILE)

    ws = wb[sheet_name]

    existing = []

    for row in ws.iter_rows(
        min_row=2,
        values_only=True
    ):

        if row[0]:

            existing.append(row)

    all_rows = existing + rows

    valid_rows = []

    for r in all_rows:

        try:

            datetime.strptime(
                r[1],
                "%Y-%m-%d"
            )

            valid_rows.append(r)

        except:
            continue

    all_rows_sorted = sorted(
        valid_rows,
        key=lambda x: datetime.strptime(
            x[1],
            "%Y-%m-%d"
        )
    )

    ws.delete_rows(2, ws.max_row)

    for r in all_rows_sorted:

        ws.append(r)

    wb.save(MASTER_FILE)

    wb.close()

# ============================================================
# PALO ALTO
# FINAL STABLE VERSION
# WORKS IN GITHUB ACTIONS + LOCAL
# ============================================================

def run_paloalto():

    print("\nRunning Palo Alto RSS Fetch...\n")

    RSS_URL = "https://security.paloaltonetworks.com/rss.xml"

    existing_keys = load_existing_keys("PALOALTO")

    new_rows = []

    # ========================================================
    # EXTRACT UPDATED DATE
    # ========================================================

    def extract_updated_date(url):

        try:

            response = requests.get(
                url,
                headers=HEADERS,
                timeout=60
            )

            html = response.text

        except Exception as e:

            print("FAILED:", url)

            print(e)

            return "UNKNOWN"

        # ====================================================
        # PRIMARY MATCH
        # ====================================================

        updated_match = re.search(
            r'Updated\s*</[^>]+>\s*<[^>]+>\s*(\d{4}-\d{2}-\d{2})',
            html,
            re.IGNORECASE
        )

        if updated_match:

            return updated_match.group(1)

        # ====================================================
        # SECONDARY MATCH
        # ====================================================

        updated_match = re.search(
            r'Updated.*?(\d{4}-\d{2}-\d{2})',
            html,
            re.IGNORECASE | re.DOTALL
        )

        if updated_match:

            return updated_match.group(1)

        # ====================================================
        # FALLBACK
        # ====================================================

        all_dates = re.findall(
            r'\d{4}-\d{2}-\d{2}',
            html
        )

        if all_dates:

            return max(all_dates)

        return "UNKNOWN"

    # ========================================================
    # PARSE RSS FEED
    # ========================================================

    feed = feedparser.parse(RSS_URL)

    for entry in feed.entries:

        title = entry.title.strip()

        link = entry.link.strip()

        # ====================================================
        # EXTRACT CVE
        # ====================================================

        cve_match = re.search(
            CVE_PATTERN,
            title,
            re.IGNORECASE
        )

        if not cve_match:
            continue

        cve = clean_cve(
            cve_match.group()
        )

        # ====================================================
        # GET UPDATED DATE
        # ====================================================

        updated_date = extract_updated_date(link)

        print("\n" + "=" * 70)

        print("CVE           :", cve)

        print("UPDATED DATE  :", updated_date)

        print("LINK          :", link)

        # ====================================================
        # DUPLICATE LOGIC
        # ====================================================

        key = (
            cve,
            updated_date,
            link
        )

        if key not in existing_keys:

            new_rows.append((
                cve,
                updated_date,
                link
            ))

            existing_keys.add(key)

    # ========================================================
    # SAVE
    # ========================================================

    if new_rows:

        append_rows(
            "PALOALTO",
            list(reversed(new_rows))
        )

        print(
            f"\nNEW ROWS ADDED : {len(new_rows)}"
        )

    else:

        print("\nNo new rows found")

# =========================================================
# BLEEPING COMPUTER
# =========================================================

def run_bleeping():

    print("\nRunning BleepingComputer...\n")

    RSS_URL = "https://www.bleepingcomputer.com/feed/"

    existing_keys = load_existing_keys(
        "BLEEPING"
    )

    new_rows = []

    feed = feedparser.parse(
        RSS_URL
    )

    for entry in feed.entries:

        if not hasattr(
            entry,
            "published_parsed"
        ):
            continue

        pub = datetime(
            *entry.published_parsed[:6]
        )

        date_str = pub.strftime(
            "%Y-%m-%d"
        )

        link = entry.link

        try:

            res = scraper.get(
                link,
                timeout=10
            )

        except:
            continue

        soup = BeautifulSoup(
            res.text,
            "html.parser"
        )

        text = soup.get_text()

        raw = re.findall(
            CVE_PATTERN,
            text,
            re.IGNORECASE
        )

        cves = {
            clean_cve(c)
            for c in raw
        }

        for c in cves:

            key = (
                c,
                date_str,
                link
            )

            if key not in existing_keys:

                new_rows.append((
                    c,
                    date_str,
                    link
                ))

                existing_keys.add(key)

    if new_rows:

        append_rows(
            "BLEEPING",
            list(reversed(new_rows))
        )


# =========================================================
# HACKERNEWS
# =========================================================

def run_hackernews():

    print("\nRunning HackerNews...\n")

    RSS_URL = (
        "https://feeds.feedburner.com/TheHackersNews"
    )

    existing_keys = load_existing_keys(
        "THN"
    )

    new_rows = []

    feed = feedparser.parse(
        RSS_URL
    )

    for entry in feed.entries:

        if not hasattr(
            entry,
            "published_parsed"
        ):
            continue

        pub = datetime(
            *entry.published_parsed[:6]
        )

        date_str = pub.strftime(
            "%Y-%m-%d"
        )

        link = entry.link

        try:

            res = requests.get(
                link,
                headers=HEADERS,
                timeout=10
            )

        except:
            continue

        soup = BeautifulSoup(
            res.text,
            "html.parser"
        )

        body = (
            soup.find(
                "div",
                class_="articlebody"
            )
            or
            soup.find(
                "div",
                class_="post-body"
            )
        )

        text = (
            body.get_text()
            if body
            else soup.get_text()
        )

        raw = re.findall(
            CVE_PATTERN,
            text,
            re.IGNORECASE
        )

        cves = {
            clean_cve(c)
            for c in raw
        }

        for c in cves:

            key = (
                c,
                date_str,
                link
            )

            if key not in existing_keys:

                new_rows.append((
                    c,
                    date_str,
                    link
                ))

                existing_keys.add(key)

    if new_rows:

        append_rows(
            "THN",
            new_rows
        )

# =========================================================
# CYBERSECURITYNEWS
# =========================================================

def run_cybersecuritynews():

    print("\nRunning CyberSecurityNews...\n")

    RSS_URL = "https://cybersecuritynews.com/feed/"

    existing_keys = load_existing_keys(
        "CYBERSECNEWS"
    )

    new_rows = []

    feed = feedparser.parse(
        RSS_URL
    )

    with sync_playwright() as p:

        browser = p.chromium.launch(
            headless=True
        )

        for entry in feed.entries:

            if not hasattr(
                entry,
                "published_parsed"
            ):
                continue

            pub = datetime(
                *entry.published_parsed[:6]
            )

            date_str = pub.strftime(
                "%Y-%m-%d"
            )

            link = entry.link

            try:

                page = browser.new_page()

                page.goto(
                    link,
                    wait_until="domcontentloaded",
                    timeout=60000
                )

                page.wait_for_timeout(3000)

                html = page.content()

                page.close()

            except:
                continue

            soup = BeautifulSoup(
                html,
                "html.parser"
            )

            article = (
                soup.find(
                    "div",
                    class_="td-post-content"
                )
                or
                soup.find(
                    "div",
                    class_="entry-content"
                )
                or
                soup.find("article")
            )

            if not article:
                continue

            text = article.get_text(
                " ",
                strip=True
            )

            raw = re.findall(
                CVE_PATTERN,
                text,
                re.IGNORECASE
            )

            cves = {
                clean_cve(c)
                for c in raw
            }

            for c in cves:

                key = (
                    c,
                    date_str,
                    link
                )

                if key not in existing_keys:

                    new_rows.append((
                        c,
                        date_str,
                        link
                    ))

                    existing_keys.add(key)

        browser.close()

    if new_rows:

        append_rows(
            "CYBERSECNEWS",
            list(reversed(new_rows))
        )


# =========================================================
# CISA
# =========================================================

def run_cisa():

    print("\nRunning CISA...\n")

    JSON_URL = (
        "https://www.cisa.gov/sites/default/files/"
        "feeds/known_exploited_vulnerabilities.json"
    )

    existing_keys = load_existing_keys(
        "CISA"
    )

    new_rows = []

    try:

        response = requests.get(
            JSON_URL,
            timeout=30
        )

        data = response.json()

    except:

        return

    vulns = data.get(
        "vulnerabilities",
        []
    )

    for item in vulns:

        cve = item.get(
            "cveID",
            ""
        ).strip()

        date_added = item.get(
            "dateAdded",
            ""
        ).strip()

        link = (
            f"https://www.cve.org/CVERecord?id={cve}"
        )

        key = (
            cve,
            date_added,
            link
        )

        if key not in existing_keys:

            new_rows.append((
                cve,
                date_added,
                link
            ))

            existing_keys.add(key)

    if new_rows:

        append_rows(
            "CISA",
            list(reversed(new_rows))
        )


# =========================================================
# SECURITYWEEK
# CLEANED VERSION
# =========================================================

def run_securityweek():

    print("\nRunning SecurityWeek...\n")

    RSS_URL = (
        "https://www.securityweek.com/category/vulnerabilities/feed/"
    )

    existing_keys = load_existing_keys(
        "SECURITYWEEK"
    )

    new_rows = []

    response = requests.get(
        RSS_URL,
        headers=HEADERS,
        timeout=30
    )

    feed = feedparser.parse(
        response.content
    )

    with sync_playwright() as playwright:

        for entry in feed.entries:

            if not hasattr(
                entry,
                "published_parsed"
            ):
                continue

            pub = datetime(
                *entry.published_parsed[:6]
            )

            date_str = pub.strftime(
                "%Y-%m-%d"
            )

            link = entry.link.strip()

            browser = playwright.chromium.launch(
                headless=True
            )

            page = browser.new_page()

            try:

                page.goto(
                    link,
                    wait_until="domcontentloaded",
                    timeout=60000
                )

                page.wait_for_timeout(4000)

                html = page.content()

            except:

                browser.close()

                continue

            browser.close()

            soup = BeautifulSoup(
                html,
                "html.parser"
            )

            article = (

                soup.find("article")

                or

                soup.find("main")

                or

                soup
            )

            # =====================================================
            # REMOVE NOISE
            # =====================================================

            for tag in article.find_all(

                [
                    "aside",
                    "script",
                    "style",
                    "noscript",
                    "iframe",
                    "footer"
                ]

            ):
                tag.decompose()

            # =====================================================
            # MAIN ARTICLE TEXT ONLY
            # =====================================================

            text_parts = []

            for p in article.find_all(["p", "li"]):

                line = p.get_text(
                    " ",
                    strip=True
                )

                if line:

                    text_parts.append(line)

            text = " ".join(text_parts)

            # =====================================================
            # EXTRACT CVEs
            # =====================================================

            raw = re.findall(
                CVE_PATTERN,
                text,
                re.IGNORECASE
            )

            cves = {
                clean_cve(c)
                for c in raw
            }

            # =====================================================
            # SAVE
            # =====================================================

            for c in cves:

                key = (
                    c,
                    date_str,
                    link
                )

                if key not in existing_keys:

                    new_rows.append((
                        c,
                        date_str,
                        link
                    ))

                    existing_keys.add(key)

            time.sleep(2)

    if new_rows:

        append_rows(
            "SECURITYWEEK",
            list(reversed(new_rows))
        )
# =========================================================
# TENABLE
# =========================================================

# =========================================================
# TENABLE
# FINAL WORKING VERSION
# =========================================================

def run_tenable():

    print("\nRunning Tenable...\n")

    URL = "https://www.tenable.com/security/research"

    existing_keys = load_existing_keys(
        "TENABLE"
    )

    new_rows = []

    with sync_playwright() as p:

        browser = p.chromium.launch(
            headless=True
        )

        page = browser.new_page()

        page.goto(
            URL,
            wait_until="domcontentloaded",
            timeout=60000
        )

        # =====================================================
        # SCROLL FULL PAGE
        # =====================================================

        previous_height = 0

        while True:

            page.evaluate(
                "window.scrollTo(0, document.body.scrollHeight)"
            )

            page.wait_for_timeout(3000)

            current_height = page.evaluate(
                "document.body.scrollHeight"
            )

            if current_height == previous_height:
                break

            previous_height = current_height

        print(
            "\nFull page loaded successfully"
        )

        html = page.content()

        browser.close()

    # =========================================================
    # PARSE HTML
    # =========================================================

    soup = BeautifulSoup(
        html,
        "html.parser"
    )

    tr_list = soup.find_all("tr")

    print("\nRows Found:", len(tr_list))

    for tr in tr_list:

        tds = tr.find_all("td")

        if len(tds) < 5:
            continue

        try:

            # =================================================
            # DATE
            # =================================================

            date_raw = tds[0].get_text(
                " ",
                strip=True
            )

            try:

                date_text = datetime.strptime(
                    date_raw,
                    "%B %d, %Y"
                ).strftime("%Y-%m-%d")

            except:

                continue

            # =================================================
            # CVE COLUMN
            # =================================================

            cve_text = tds[4].get_text(
                " ",
                strip=True
            )

            # =================================================
            # EXTRACT CVEs
            # =================================================

            found_cves = re.findall(
                CVE_PATTERN,
                cve_text,
                re.IGNORECASE
            )

            if not found_cves:
                continue

            # =================================================
            # LINK
            # =================================================

            link = URL

            # =================================================
            # SAVE
            # =================================================

            for cve in found_cves:

                cve = cve.upper().strip()

                key = (
                    cve,
                    date_text,
                    link
                )

                if key in existing_keys:
                    continue

                print(
                    "ADDING:",
                    cve,
                    date_text
                )

                new_rows.append((
                    cve,
                    date_text,
                    link
                ))

                existing_keys.add(key)

        except Exception as e:

            print(
                "\nERROR PARSING ROW"
            )

            print(e)

    # =========================================================
    # SAVE
    # =========================================================

    if new_rows:

        append_rows(
            "TENABLE",
            list(reversed(new_rows))
        )
# =========================================================
# ZDI
# =========================================================

def run_zdi():

    print("\nRunning ZDI...\n")

    RSS_URL = (
        "https://www.zerodayinitiative.com/rss/published/"
    )

    existing_keys = load_existing_keys(
        "ZDI"
    )

    new_rows = []

    feed = feedparser.parse(
        RSS_URL
    )

    for entry in feed.entries:

        if not hasattr(
            entry,
            "published_parsed"
        ):
            continue

        pub = datetime(
            *entry.published_parsed[:6]
        )

        date_str = pub.strftime(
            "%Y-%m-%d"
        )

        link = entry.link.strip()

        try:

            response = requests.get(
                link,
                headers=HEADERS,
                timeout=20
            )

        except:
            continue

        soup = BeautifulSoup(
            response.text,
            "html.parser"
        )

        article = (
            soup.find(
                "div",
                class_="container"
            )
            or
            soup.find("article")
            or
            soup
        )

        text = article.get_text(
            " ",
            strip=True
        )

        raw = re.findall(
            CVE_PATTERN,
            text,
            re.IGNORECASE
        )

        cves = {
            clean_cve(c)
            for c in raw
        }

        for c in cves:

            key = (
                c,
                date_str,
                link
            )

            if key not in existing_keys:

                new_rows.append((
                    c,
                    date_str,
                    link
                ))

                existing_keys.add(key)

    if new_rows:

        append_rows(
            "ZDI",
            list(reversed(new_rows))
        )


# =========================================================
# RAPID7
# FINAL WORKING VERSION
# =========================================================

def run_rapid7():

    print("\nRunning Rapid7...\n")

    BASE_URL = "https://www.rapid7.com/db/"

    existing_keys = load_existing_keys(
        "RAPID7"
    )

    new_rows = []

    # =========================================
    # SCRAPE PAGE
    # =========================================

    def get_html(playwright):

        browser = playwright.chromium.launch(
            headless=True
        )

        page = browser.new_page()

        page.goto(
            BASE_URL,
            wait_until="domcontentloaded",
            timeout=120000
        )

        page.wait_for_timeout(5000)

        while True:

            try:

                load_more = page.locator(
                    "text=Load more"
                )

                if load_more.count() > 0:

                    print(
                        "Clicking Load More..."
                    )

                    load_more.first.click()

                    page.wait_for_timeout(3000)

                else:

                    break

            except:

                break

        html = page.content()

        browser.close()

        return html

    # =========================================
    # GET HTML
    # =========================================

    with sync_playwright() as playwright:

        html = get_html(playwright)

    soup = BeautifulSoup(
        html,
        "html.parser"
    )

    cards = soup.find_all(
        ["a", "div", "article"]
    )

    print("\nProcessing cards...\n")

    for card in cards:

        text = card.get_text(
            " ",
            strip=True
        )

        cves = re.findall(
            CVE_PATTERN,
            text,
            re.IGNORECASE
        )

        if not cves:
            continue

        # =====================================
        # DATE
        # =====================================

        date_match = re.search(
            r"Published:\s*([0-9]{4}-[0-9]{2}-[0-9]{2})",
            text
        )

        if not date_match:
            continue

        publish_date_raw = date_match.group(1)

        try:

            publish_date = datetime.strptime(
                publish_date_raw,
                "%Y-%m-%d"
            ).strftime("%Y-%m-%d")

        except:

            continue

        # =====================================
        # LINK
        # =====================================

        href = card.get("href")

        if href:

            if href.startswith("/"):

                link = (
                    "https://www.rapid7.com"
                    + href
                )

            else:

                link = href

        else:

            link = BASE_URL

        # =====================================
        # SAVE CVEs
        # =====================================

        for cve in cves:

            cve = cve.upper()

            key = (
                cve,
                publish_date,
                link
            )

            if key in existing_keys:
                continue

            print(
                "ADDING:",
                cve,
                publish_date
            )

            new_rows.append((
                cve,
                publish_date,
                link
            ))

            existing_keys.add(key)

    # =========================================
    # SAVE
    # =========================================

    if new_rows:

        append_rows(
            "RAPID7",
            list(reversed(new_rows))
        )


# =========================================================
# ZERODAYCZ
# ORIGINAL WORKING VERSION
# =========================================================

def run_zerodaycz():

    print("\nRunning Zero-Day.cz...\n")

    URL = "https://www.zero-day.cz/database/"

    existing_keys = load_existing_keys(
        "ZERODAYCZ"
    )

    new_rows = []

    processed = set()

    with sync_playwright() as p:

        browser = p.chromium.launch(
            headless=True
        )

        page = browser.new_page()

        page.goto(
            URL,
            wait_until="domcontentloaded",
            timeout=120000
        )

        previous_height = 0

        while True:

            page.evaluate(
                "window.scrollTo(0, document.body.scrollHeight)"
            )

            page.wait_for_timeout(3000)

            current_height = page.evaluate(
                "document.body.scrollHeight"
            )

            if current_height == previous_height:
                break

            previous_height = current_height

        print(
            "\nFull page loaded successfully"
        )

        html = page.content()

        browser.close()

    soup = BeautifulSoup(
        html,
        "html.parser"
    )

    all_blocks = soup.find_all(

        "div",

        class_=lambda x:
        x and "col" in str(x)
    )

    print("\nParsing CVEs...\n")

    for block in all_blocks:

        text = block.get_text(
            " ",
            strip=True
        )

        cve_match = re.search(
            r"CVE-\d{4}-\d{4,7}",
            text,
            re.IGNORECASE
        )

        if not cve_match:
            continue

        cve = cve_match.group(0).upper()

        title_tag = (

            block.find("h1")

            or

            block.find("h2")

            or

            block.find("h3")

            or

            block.find("h4")
        )

        if title_tag:

            title = title_tag.get_text(
                " ",
                strip=True
            )

        else:

            title = cve

        date_matches = re.findall(
            r"\d{4}-\d{2}-\d{2}",
            text
        )

        if not date_matches:
            continue

        publish_date = date_matches[0]

        unique_card = (
            cve,
            publish_date,
            title
        )

        if unique_card in processed:
            continue

        processed.add(unique_card)

        link = URL

        key = (
            cve,
            publish_date,
            link
        )

        if key in existing_keys:
            continue

        existing_keys.add(key)

        print("\n" + "=" * 70)

        print("TITLE :", title)

        print("CVE   :", cve)

        print("DATE  :", publish_date)

        new_rows.append((
            cve,
            publish_date,
            link
        ))

    # =====================================================
    # SAVE
    # =====================================================

    if new_rows:

        append_rows(
            "ZERODAYCZ",
            list(reversed(new_rows))
        )
# =========================================================
# MAIN
# =========================================================

def main():

    create_workbook_if_missing()

    run_paloalto()
    
    run_bleeping()

    run_hackernews()

    run_cybersecuritynews()

    run_cisa()

    run_securityweek()

    run_tenable()

    

    run_rapid7()

    run_zerodaycz()

    print("\nALL SOURCES COMPLETED")

# =========================================================
# RUN
# =========================================================

if __name__ == "__main__":

    main()

